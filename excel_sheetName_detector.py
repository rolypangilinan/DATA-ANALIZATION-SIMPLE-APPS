# import os
# import pandas as pd
# from pathlib import Path

# # Define the folder path (use raw string or double backslashes for Windows network paths)
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Target sheet name (exact match)
# TARGET_SHEET = "in-process&performance NG"

# def check_sheet_exists(file_path, sheet_name):
#     """
#     Check if the specified sheet exists in the Excel file.
#     Returns True if exists, False otherwise.
#     """
#     try:
#         # Get list of sheet names
#         xl = pd.ExcelFile(file_path)
#         return sheet_name in xl.sheet_names
#     except Exception as e:
#         print(f"Error reading {file_path}: {e}")
#         return False

# def main():
#     # Check if folder exists and is accessible
#     if not os.path.exists(FOLDER_PATH):
#         print(f"Folder not found or inaccessible: {FOLDER_PATH}")
#         return

#     # Get all .xlsx files in the folder (non-recursive)
#     excel_files = [f for f in os.listdir(FOLDER_PATH) if f.lower().endswith('.xlsx')]

#     if not excel_files:
#         print("No Excel (.xlsx) files found in the folder.")
#         return

#     print(f"Scanning {len(excel_files)} Excel file(s) in: {FOLDER_PATH}\n")

#     # Process each file
#     for i, filename in enumerate(sorted(excel_files), 1):
#         full_path = os.path.join(FOLDER_PATH, filename)
        
#         # Check if the target sheet exists
#         exists = check_sheet_exists(full_path, TARGET_SHEET)
        
#         # Format number with leading zero
#         num_str = f"{i:02d}"
        
#         # Print in requested format
#         status = "YES" if exists else "NO"
#         print(f"{num_str}. {filename} {status}")

# if __name__ == "__main__":
#     main()
















# #%%

# import os
# import pandas as pd

# # Define the folder path (use raw string for Windows network paths)
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Target sheet we are looking for
# TARGET_SHEET = "in-process&performance NG"

# # Reference sheet whose name we want to show when target is missing
# REFERENCE_SHEET_NAME = "R-HISTORY MASTER FILE"

# def get_sheet_info(file_path):
#     """
#     Returns:
#     - bool: True if TARGET_SHEET exists
#     - str: Name of the sheet that matches REFERENCE_SHEET_NAME (or None)
#     """
#     try:
#         xl = pd.ExcelFile(file_path)
#         sheet_names = xl.sheet_names
        
#         target_exists = TARGET_SHEET in sheet_names
        
#         # Find the sheet that matches the reference name
#         reference_sheet = None
#         for sheet in sheet_names:
#             if sheet.strip().upper() == REFERENCE_SHEET_NAME.strip().upper():
#                 reference_sheet = sheet  # preserve original casing
#                 break
        
#         return target_exists, reference_sheet
    
#     except Exception as e:
#         print(f"Error reading {file_path}: {e}")
#         return False, None

# def main():
#     if not os.path.exists(FOLDER_PATH):
#         print(f"Folder not found or inaccessible: {FOLDER_PATH}")
#         return

#     # Get all .xlsx files
#     excel_files = [f for f in os.listdir(FOLDER_PATH) if f.lower().endswith('.xlsx')]

#     if not excel_files:
#         print("No Excel (.xlsx) files found in the folder.")
#         return

#     print(f"Scanning {len(excel_files)} Excel file(s) in: {FOLDER_PATH}\n")

#     for i, filename in enumerate(sorted(excel_files), 1):
#         full_path = os.path.join(FOLDER_PATH, filename)
        
#         target_exists, reference_sheet_name = get_sheet_info(full_path)
        
#         num_str = f"{i:02d}"
        
#         if target_exists:
#             print(f"{num_str}. {filename} YES")
#         else:
#             extra_info = f" (R-HISTORY MASTER FILE: {reference_sheet_name})" if reference_sheet_name else ""
#             print(f"{num_str}. {filename} NO{extra_info}")

# if __name__ == "__main__":
#     main()










# #%%
# import os
# import pandas as pd
# from pathlib import Path

# # Define the folder path
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Target sheet name to look for (case-insensitive)
# TARGET_SHEET_PART = "in-process&performance NG".lower()

# def check_sheet_exists(file_path):
#     try:
#         # Read only the sheet names
#         xl = pd.ExcelFile(file_path)
        
#         # Check if there is at least 2 sheets
#         if len(xl.sheet_names) < 2:
#             return False
        
#         # Get the name of the 2nd sheet (index 1)
#         second_sheet_name = xl.sheet_names[1]
        
#         # Check if target string is in the second sheet name (case-insensitive)
#         return TARGET_SHEET_PART in second_sheet_name.lower()
    
#     except Exception as e:
#         # If file can't be opened (corrupt, not Excel, permission issue, etc.)
#         print(f"  Warning: Could not read {file_path.name} - {e}")
#         return False

# def main():
#     print(f"Scanning folder: {FOLDER_PATH}\n")
    
#     # Get all .xlsx files in the folder (non-recursive)
#     excel_files = sorted(Path(FOLDER_PATH).glob("*.xlsx"))
    
#     if not excel_files:
#         print("No .xlsx files found in the folder.")
#         return
    
#     print("FILE NAME                  | YES/NO")
#     print("-" * 45)
    
#     for i, file_path in enumerate(excel_files, 1):
#         # Get just the filename
#         filename = file_path.name
        
#         # Check if the second sheet contains the target name
#         exists = check_sheet_exists(file_path)
        
#         # Format output with number and month-like style
#         status = "YES" if exists else "NO"
#         print(f"{i:02d}. {filename:<30} {status}")
    
#     print("\nDone.")

# if __name__ == "__main__":
#     main()











# #%%
# import os
# import pandas as pd

# # Define the network path to the Excel file
# file_path = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Check if the file exists
# if not os.path.exists(file_path):
#     print(f"Error: File not found at {file_path}")
#     exit()

# # Get just the file name for display
# file_name = os.path.basename(file_path)

# # Read the Excel file and get all sheet names
# try:
#     excel_file = pd.ExcelFile(file_path)
#     sheet_names = excel_file.sheet_names
# except Exception as e:
#     print(f"Error reading the Excel file: {e}")
#     exit()

# # Print the output in the requested format
# print(file_name)
# print()  # empty line after file name

# for i, sheet in enumerate(sheet_names, 1):
#     # Format number with leading zero if < 10
#     num_str = f"{i:02d}" if i < 10 else f"{i}"
#     print(f"{num_str}. {sheet}")

# # Optional: show total number of sheets
# print(f"\nTotal sheets: {len(sheet_names)}")










# #%%

# import os
# from openpyxl import load_workbook

# # IMPORTANT: Make sure you have read access to this network path
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# def main():
#     # Check if folder is accessible
#     if not os.path.exists(FOLDER_PATH):
#         print(f"Error: Folder not found or inaccessible:\n{FOLDER_PATH}")
#         return

#     # Get all Excel files and sort them
#     excel_files = [f for f in os.listdir(FOLDER_PATH)
#                    if f.lower().endswith(('.xlsx', '.xlsm', '.xls'))]
#     excel_files.sort()

#     if not excel_files:
#         print("No Excel files found in the folder.")
#         return

#     # Header
#     print(f"{'FILE NAME':<45} 2ND SHEET NAME")
#     print("-" * 70)

#     for filename in excel_files:
#         full_path = os.path.join(FOLDER_PATH, filename)
#         try:
#             # Load workbook in read-only mode (faster & safer)
#             wb = load_workbook(full_path, read_only=True, data_only=True)
#             sheet_names = wb.sheetnames

#             if len(sheet_names) >= 2:
#                 second_sheet = sheet_names[1]
#             else:
#                 second_sheet = "NO SECOND SHEET"

#             # Print with aligned formatting
#             print(f"{filename:<45} {second_sheet}")

#             wb.close()

#         except Exception as e:
#             print(f"{filename:<45} ERROR: {str(e)}")

# if __name__ == "__main__":
#     main()












# #%%
# # TESTED BUT HAS AN ERROR

# import os
# import openpyxl
# from pathlib import Path

# # Define the folder path
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Keyword to search for (case-insensitive)
# KEYWORD = "performance"

# def contains_performance(sheet_name):
#     """Check if 'performance' is in the sheet name (case-insensitive)"""
#     return KEYWORD.lower() in sheet_name.lower()

# def main():
#     print(f"Scanning folder: {FOLDER_PATH}\n")
#     print(f"Looking for sheet names containing: '{KEYWORD}'\n")
#     print("Results (files with at least one matching sheet):\n")
#     print(f"{'FILE NAME':<40} {'SHEET NAME':<35}")
#     print("-" * 80)

#     found_any = False

#     # Get all Excel files (.xlsx, .xlsm, .xls)
#     for file_path in Path(FOLDER_PATH).glob("*.[xX][lL][sS][xXmM]"):
#         file_name = file_path.name
        
#         try:
#             wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
#             matching_sheets = []
#             for sheet in wb.sheetnames:
#                 if contains_performance(sheet):
#                     matching_sheets.append(sheet)
            
#             if matching_sheets:
#                 found_any = True
#                 # Print file name only once, then all matching sheets
#                 print(f"{file_name:<40} {matching_sheets[0]:<35}")
#                 for extra_sheet in matching_sheets[1:]:
#                     print(f"{'':<40} {extra_sheet:<35}")
#                 print()  # empty line between files

#             wb.close()

#         except Exception as e:
#             print(f"Error reading {file_name}: {e}")
#             print()

#     if not found_any:
#         print("No files found with sheet names containing 'performance'.")

# if __name__ == "__main__":
#     main()


# # OUTPUT
# # FILE NAME                                SHEET NAME
# # --------------------------------------------------------------------------------
# # 01. January.xlsx                         IN-PROCESS&PERFORMANCE NG

# # 02. February.xlsx                        IN-PROCESS&PERFORMANCE NG

# # 03. March.xlsx                           IN-PROCESS&PERFORMANCE NG

# # 04. April.xlsx                           IN-PROCESS&PERFORMANCE NG

# # 05. May.xlsx                             IN-PROCESS&PERFORMANCE NG

# # 06. June.xlsx                            IN-PROCESS&PERFORMANCE NG

# # 07. July.xlsx                            IN-PROCESS&PERFORMANCE NG

# # 08. August.xlsx                          IN-PROCESS&PERFORMANCE NG

# # 09. September.xlsx                       IN-PROCESS&PERFORMANCE NG

# # 10. October.xlsx                         IN-PROCESS&PERFORMANCE NG

# # 11.November.xlsx                         IN-PROCESS&PERFORMANCE NG

# # 12. December.xlsx                        IN-PROCESS&PERFORMANCE NG


























#%%

# # TESTED AND NO ERROR

# import os
# import openpyxl
# from pathlib import Path

# # Define the folder path
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Keyword to search for (case-insensitive)
# KEYWORD = "performance"

# def contains_performance(sheet_name):
#     """Check if 'performance' is in the sheet name (case-insensitive)"""
#     return KEYWORD.lower() in sheet_name.lower()

# def main():
#     print(f"Scanning folder: {FOLDER_PATH}\n")
#     print(f"Looking for sheet names containing: '{KEYWORD}'\n")
#     print("Results (files with at least one matching sheet):\n")
#     print(f"{'FILE NAME':<40} {'SHEET NAME':<35}")
#     print("-" * 80)

#     found_any = False

#     # Get all Excel files (.xlsx, .xlsm, .xls)
#     for file_path in Path(FOLDER_PATH).glob("*.[xX][lL][sS][xXmM]"):
#         file_name = file_path.name
        
#         # Skip temporary hidden files starting with ~$
#         if file_name.startswith('~$'):
#             continue
        
#         try:
#             wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
#             matching_sheets = []
#             for sheet in wb.sheetnames:
#                 if contains_performance(sheet):
#                     matching_sheets.append(sheet)
            
#             if matching_sheets:
#                 found_any = True
#                 # Print file name only once, then all matching sheets
#                 print(f"{file_name:<40} {matching_sheets[0]:<35}")
#                 for extra_sheet in matching_sheets[1:]:
#                     print(f"{'':<40} {extra_sheet:<35}")
#                 print()  # empty line between files

#             wb.close()

#         except Exception as e:
#             print(f"Error reading {file_name}: {e}")
#             print()

#     if not found_any:
#         print("No files found with sheet names containing 'performance'.")

# if __name__ == "__main__":
#     main()




















#%%
# TESTED AND NO ERROR
# COUNTS HOW MANY EXCEL FILES HAVE EACH SHEET NAME CONTAINING "PERFORMANCE"

# import os
# import openpyxl
# from pathlib import Path
# from collections import defaultdict

# # Define the folder path
# FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# # Keyword to search for (case-insensitive)
# KEYWORD = "performance"

# def contains_performance(sheet_name):
#     """Check if 'performance' is in the sheet name (case-insensitive)"""
#     return KEYWORD.lower() in sheet_name.lower()

# def main():
#     print(f"Scanning folder: {FOLDER_PATH}\n")
#     print(f"Looking for sheet names containing: '{KEYWORD}'\n")
    
#     # Dictionary to count how many files contain each matching sheet name
#     sheet_count = defaultdict(int)
#     # List to store (file_name, sheet_name) for detailed listing
#     file_sheet_list = []

#     found_any = False

#     # Get all Excel files (.xlsx, .xlsm, .xls)
#     for file_path in Path(FOLDER_PATH).glob("*.[xX][lL][sS][xXmM]"):
#         file_name = file_path.name
        
#         # Skip temporary hidden files starting with ~$
#         if file_name.startswith('~$'):
#             continue
        
#         try:
#             wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
#             matching_sheets = []
#             for sheet in wb.sheetnames:
#                 if contains_performance(sheet):
#                     matching_sheets.append(sheet)
            
#             if matching_sheets:
#                 found_any = True
#                 for sheet in matching_sheets:
#                     sheet_count[sheet] += 1
#                     file_sheet_list.append((file_name, sheet))

#             wb.close()

#         except Exception as e:
#             print(f"Error reading {file_name}: {e}")
#             print()

#     # === DISPLAY RESULTS ===

#     if not found_any:
#         print("No files found with sheet names containing 'performance'.")
#         return

#     # First: Detailed list of files and their matching sheets
#     print("Results (files with at least one matching sheet):\n")
#     print(f"{'FILE NAME':<40} {'SHEET NAME':<35}")
#     print("-" * 80)

#     for file_name, sheet_name in file_sheet_list:
#         print(f"{file_name:<40} {sheet_name:<35}")

#     print("\n" + "="*80 + "\n")

#     # Second: Summary - how many files have each sheet name
#     print("SUMMARY: Count of files containing each sheet name\n")
#     print(f"{'SHEET NAME':<40} {'NUMBER OF FILES'}")
#     print("-" * 60)

#     # Sort by count descending, then by sheet name
#     for sheet, count in sorted(sheet_count.items(), key=lambda x: (-x[1], x[0])):
#         print(f"{sheet:<40} {count}")

#     print(f"\nTotal unique sheet names found: {len(sheet_count)}")
#     print(f"Total Excel files processed with matching sheets: {len(file_sheet_list)}")

# if __name__ == "__main__":
#     main()
































#%%

# TESTED

import os
import openpyxl
from pathlib import Path
from collections import defaultdict

# Define the folder path
FOLDER_PATH = r"\\192.168.2.19\production\DIS-ASSY. HISTORY OF 2025"

# Keyword to search for (case-insensitive)
KEYWORD = "performance"

def contains_performance(sheet_name):
    """Check if 'performance' is in the sheet name (case-insensitive)"""
    return KEYWORD.lower() in sheet_name.lower()

def main():
    print(f"Scanning folder: {FOLDER_PATH}\n")
    print(f"Looking for sheet names containing: '{KEYWORD}'\n")
    
    # Dictionary to count how many files contain each matching sheet name
    sheet_count = defaultdict(int)
    # List to store (file_name, sheet_name) for detailed listing of files WITH the keyword
    files_with_sheet = []
    # List to store file names that have NO matching sheet
    files_without_sheet = []

    found_any = False

    # Get all Excel files (.xlsx, .xlsm, .xls)
    for file_path in Path(FOLDER_PATH).glob("*.[xX][lL][sS][xXmM]"):
        file_name = file_path.name
        
        # Skip temporary hidden files starting with ~$
        if file_name.startswith('~$'):
            continue
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            matching_sheets = []
            for sheet in wb.sheetnames:
                if contains_performance(sheet):
                    matching_sheets.append(sheet)
            
            if matching_sheets:
                found_any = True
                for sheet in matching_sheets:
                    sheet_count[sheet] += 1
                    files_with_sheet.append((file_name, sheet))
            else:
                # No matching sheet found in this file
                files_without_sheet.append(file_name)

            wb.close()

        except Exception as e:
            print(f"Error reading {file_name}: {e}")
            print()
            # Even if error, we can still count it as "without" if we want, but safer to skip detailed listing
            files_without_sheet.append(file_name)

    # === DISPLAY RESULTS ===

    # 1. Files WITH matching sheets
    if files_with_sheet:
        print("FILES CONTAINING SHEET NAMES WITH 'performance':\n")
        print(f"{'FILE NAME':<40} {'SHEET NAME':<35}")
        print("-" * 80)
        for file_name, sheet_name in files_with_sheet:
            print(f"{file_name:<40} {sheet_name:<35}")
        print(f"\nTotal files with matching sheets: {len(set(file for file, _ in files_with_sheet))}")
    else:
        print("No files found with sheet names containing 'performance'.")

    print("\n" + "="*80 + "\n")

    # 2. Summary of sheet name occurrences
    if sheet_count:
        print("SUMMARY: Count of files containing each sheet name\n")
        print(f"{'SHEET NAME':<40} {'NUMBER OF FILES'}")
        print("-" * 60)
        for sheet, count in sorted(sheet_count.items(), key=lambda x: (-x[1], x[0])):
            print(f"{sheet:<40} {count}")
        print(f"\nTotal unique sheet names found: {len(sheet_count)}")
        print(f"Total occurrences across files: {sum(sheet_count.values())}")
    else:
        print("SUMMARY: No matching sheet names found in any file.\n")

    print("="*80 + "\n")

    # 3. Files WITHOUT any matching sheet
    if files_without_sheet:
        print("FILES WITHOUT ANY SHEET NAME CONTAINING 'performance':\n")
        print(f"{'FILE NAME':<50}")
        print("-" * 50)
        for file_name in sorted(files_without_sheet):
            print(f"{file_name:<50}")
        print(f"\nTotal files without matching sheets: {len(files_without_sheet)}")
    else:
        print("All processed Excel files contain at least one sheet with 'performance'.")

    # Final overall stats
    total_processed = len(files_with_sheet) + len(files_without_sheet) - len([f for f in files_without_sheet if 'Error' in str(f)])  # rough adjustment
    print(f"\nOverall: {len(set(file for file, _ in files_with_sheet)) + len([f for f in files_without_sheet if f not in [e[0] for e in files_with_sheet]])} valid Excel files scanned.")

if __name__ == "__main__":
    main()